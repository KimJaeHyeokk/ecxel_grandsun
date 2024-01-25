from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import os


def auto_size_columns(sheet, min_width=8, max_padding=15):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = min(max_length + max_padding, min_width)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        except:
            pass
def base_sheet(ws):
    
    ws.merge_cells('B2:B5')
    ws.merge_cells('M2:O5')
    ws.merge_cells('Q2:AB2')
    ws.merge_cells('AD2:BA2')

    ws.merge_cells('D2:K2')
    ws.merge_cells('D3:H3')
    ws.merge_cells('I3:J3')
    ws.merge_cells('K3:K4')
    ws.merge_cells('B6:B29')
    ws['AD2'] = '인버터 출력'
    ws['B2'] = '일자'
    ws['M2'] = '실증지 일간 합계'
    ws['M6'] = '현재발전 합계(kW)'
    ws['N6'] = '일간발전량 합계(kWh)'
    ws['O6'] = '누적발전량 합계(MWh)'
    ws['Q2'] = '인버터 입력'
    ws['D5'] = 'W/㎡'
    ws['E5'] = 'W/㎡'
    ws['F5'] = 'W/㎡'
    ws['G5'] = 'W/㎡'
    ws['H5'] = 'W/㎡'

    ws['D2'] = '기상 데이터'
    ws['D3'] = '일사량'
    ws['I3'] = '온도'
    ws['K3'] = '풍속'
    ws['K5'] = 'm/s'
    ws['I4'] = '외기온도'
    ws['J4'] = '모듈온도'
    ws['I5'] = '℃'
    ws['J5'] = '℃'
    # 데이터 입력
    time_columns = ['C', 'L', 'P','AC']

    for col in time_columns:
        # '시간' 셀 입력
        merge_range = f'{col}2:{col}5'
        ws.merge_cells(merge_range)
        ws[f'{col}2'] = '시간'
    # '0:00'부터 '23:00'까지 시간 입력
        start_time = datetime.strptime('00:00', '%H:%M')
        for i in range(6, 30):
            time_str = (start_time + timedelta(hours=i - 6)).strftime('%H:%M')
            ws.cell(row=i, column=ws[col + '2'].column, value=time_str)
    # 전체 셀 가운데 정렬 설정
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2,min_col=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            cell.border = border
    # 모든 열의 너비 자동 조절
    auto_size_columns(ws)
    ws.column_dimensions['B'].width = 100

excel_date = datetime.now()
year = excel_date.strftime('%Y')+'년'
month = excel_date.strftime('%m')+'월'

base_folder_path = 'C:\Rnd_report'

year_folder_path = os.path.join(base_folder_path, str(year))
month_folder_path = os.path.join(year_folder_path, month)
print("month_folder_path",month_folder_path)

# folder_path = 'C:\grandsun_excel'

if not os.path.exists(month_folder_path):
    os.makedirs(month_folder_path)
    print(f"Folder '{month_folder_path}' created successfully!")

wb = Workbook()
ws = wb.active
now_date = datetime(datetime.today().year,datetime.today().month,datetime.today().day, datetime.today().hour, datetime.today().minute, datetime.today().second)
day_date = now_date.strftime("%Y-%m-%d")
date_id = int(day_date.replace("-",""))
ws.title = 'demon_area_name' #첫번째 자동으로 만들어지는 sheet
second_sheet = wb.create_sheet("demon_area_name") # 새롭게 만들어지는 시트 이름 정하기     #근데 이쪽에는 값 어케느냐..
base_sheet(ws)
base_sheet(second_sheet)
# 셀 테두리 설정
wb.save(os.path.join(month_folder_path, f'일간 보고서_test({date_id}).xlsx'))